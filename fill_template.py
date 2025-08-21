# fill_template.py
import re
from pathlib import Path
import openpyxl

def find_reports_path() -> Path:
    root = Path(".")
    candidates = list(root.glob("All_Reports*.xlsx"))
    if candidates:
        return candidates[0]
    for p in root.glob("*.xlsx"):
        if "Rekenmodel Vivada" not in p.name:
            return p
    raise FileNotFoundError("All_Reports.xlsx niet gevonden in repo root")

def find_template_path() -> Path:
    root = Path(".")
    candidates = list(root.glob("Rekenmodel Vivada*.xlsx"))
    if not candidates:
        raise FileNotFoundError("Vivada template niet gevonden in repo root")
    return candidates[0]

cell_map = {
    "address": "C7",
    "city": "C8",
    "country": "C9",
    "object_type": "C10",
    "lettable_floor_area_sqm": "F12",
    "cadastral_surface_sqm": "F13",
    "construction_year": "C11",
    "energy_label": "C12",
    "valuation_date": "F7",
    "appraiser": "F8",
    "valuation_approach": "C13",
    "rental_income_eur": "F16",
    "market_rent_eur": "F17",
    "tenant_lfa_sqm": "F18",
    "operating_expenses_eur": "F19",
}

def normalize_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

def read_section_table(ws, title_text: str) -> dict:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and title_text.lower() in v.lower():
            data = {}
            rr = r + 1
            while rr <= ws.max_row:
                k = ws.cell(rr, 1).value
                val = ws.cell(rr, 2).value
                if k is None and val is None:
                    break
                if k is not None:
                    data[str(k)] = val
                rr += 1
            return data
    return {}

def parse_objectinformatie_to_std(meta_raw: dict) -> dict:
    text_map = {
        "address": ["Adres", "Address"],
        "city": ["Plaats", "City"],
        "country": ["Land", "Country"],
        "object_type": ["Objecttype", "Object type"],
        "lettable_floor_area_sqm": ["VVO", "LFA", "Lettable floor area"],
        "cadastral_surface_sqm": ["Kadastrale oppervlakte", "Cadastral surface"],
        "construction_year": ["Bouwjaar", "Construction year"],
        "energy_label": ["Energielabel", "Energy label"],
        "valuation_date": ["Taxatiedatum", "Valuation date"],
        "appraiser": ["Taxateur", "Appraiser"],
        "valuation_approach": ["Waarderingsmethode", "Valuation approach", "Valuation type"],
    }
    out = {}
    for std_key, candidates in text_map.items():
        for cand in candidates:
            for raw_k, raw_v in meta_raw.items():
                if normalize_key(cand) in normalize_key(raw_k):
                    out[std_key] = raw_v
                    break
            if std_key in out:
                break
    return out

def extract_payload_from_report_sheet(ws_rep):
    meta_raw = read_section_table(ws_rep, "Objectinformatie")
    current_raw = read_section_table(ws_rep, "Current State")

    meta_std = parse_objectinformatie_to_std(meta_raw)

    cs_map = {
        "rental_income_eur": ["Rental income", "Contracthuur jr", "Contracthuur", "Huurprijs per jaar"],
        "market_rent_eur": ["Market rent", "Markthuur jr", "Markthuur"],
        "tenant_lfa_sqm": ["VVO", "GBO", "LFA"],
        "operating_expenses_eur": ["Operating expenses", "Exploitatiekosten"],
    }
    cs_std = {}
    for std_key, candidates in cs_map.items():
        for cand in candidates:
            for raw_k, raw_v in current_raw.items():
                if normalize_key(cand) in normalize_key(raw_k):
                    cs_std[std_key] = raw_v
                    break
            if std_key in cs_std:
                break

    return {**meta_std, **cs_std}

def write_value_if_input(ws, cell_addr: str, value):
    cell = ws[cell_addr]
    if isinstance(cell.value, str) and cell.value.startswith("="):
        return False
    ws[cell_addr] = value
    return True

def apply_payload_to_vivada(wb_tmpl, payload: dict, summary_sheet_name="Summary"):
    ws = wb_tmpl[summary_sheet_name] if summary_sheet_name in wb_tmpl.sheetnames else wb_tmpl.active
    for key, addr in cell_map.items():
        val = payload.get(key, None)
        if val in (None, ""):
            continue
        try:
            write_value_if_input(ws, addr, val)
        except Exception:
            pass

def main():
    reports_path = find_reports_path()
    template_path = find_template_path()
    print(f"Rapportbestand: {reports_path.name}")
    print(f"Templatebestand: {template_path.name}")

    rep_wb = openpyxl.load_workbook(reports_path, data_only=True)
    out_dir = Path("out_vivada")
    out_dir.mkdir(parents=True, exist_ok=True)

    for sheet_name in rep_wb.sheetnames:
        print(f"Verwerk sheet: {sheet_name}")
        ws_rep = rep_wb[sheet_name]
        payload = extract_payload_from_report_sheet(ws_rep)
        if not any(k in payload and payload[k] not in (None, "") for k in cell_map.keys()):
            print("Geen relevante velden, overslaan")
            continue
        wb_tmpl = openpyxl.load_workbook(template_path, data_only=False)
        apply_payload_to_vivada(wb_tmpl, payload, summary_sheet_name="Summary")
        if "Sheet" in wb_tmpl.sheetnames:
            try:
                del wb_tmpl["Sheet"]
            except Exception:
                pass
        safe_name =_

